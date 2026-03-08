from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple
from urllib.parse import quote_plus, urlparse

import pandas as pd
from apify import Actor
from playwright.async_api import async_playwright, Browser, BrowserContext, Page

from salesforce_appexchange_engine.scraper_apps import APPX_BASE, discover_app_urls, extract_app_detail


# -----------------------------
# helpers
# -----------------------------
def _as_list_str(v: Any) -> List[str]:
    if v is None:
        return []
    if isinstance(v, list):
        return [str(x).strip().lower() for x in v if str(x).strip()]
    if isinstance(v, str):
        return [v.strip().lower()] if v.strip() else []
    return []


def _to_int(v: Any, default: int) -> int:
    try:
        if v is None or v == "":
            return default
        return int(v)
    except Exception:
        return default


def _to_float(v: Any, default: float) -> float:
    try:
        if v is None or v == "":
            return default
        return float(v)
    except Exception:
        return default


def _get_hierarchy_context(config: Dict[str, Any]) -> Dict[str, Any]:
    category_group = (config.get("categoryGroup") or "").strip().lower()
    sphere = (config.get("sphere") or "").strip().lower()

    resolved = config.get("resolvedAppTypes") or []
    app_types = _as_list_str(config.get("appTypes"))

    app_type_id = ""
    app_type_title = ""

    if isinstance(resolved, list) and resolved and isinstance(resolved[0], dict):
        first = resolved[0]
        app_type_id = str(first.get("id") or "").strip().lower()
        app_type_title = str(first.get("title") or "").strip()
    elif app_types:
        app_type_id = app_types[0]
        app_type_title = app_types[0]

    return {
        "category_group": category_group,
        "sphere": sphere,
        "app_type_id": app_type_id,
        "app_type_title": app_type_title,
    }


def _score_business_needs(record: Dict[str, Any], needs: List[str]) -> Tuple[int, List[str]]:
    if not needs:
        return 0, []

    blob = " ".join(
        [
            str(record.get("primary_category_name", "")),
            str(record.get("app_name", "")),

            str(record.get("short_description", "")),
            str(record.get("app_type_title", "")),
            str(record.get("price_text", "")),
            " ".join([str(x) for x in (record.get("clouds") or [])]),
        ]
    ).lower()

    matches: List[str] = []
    score = 0
    for n in needs:
        n = (n or "").strip().lower()
        if not n:
            continue
        if n in blob:
            score += 1
            matches.append(n)

    return score, matches


def _proxy_for_playwright(config: Dict[str, Any]) -> Dict[str, Any] | None:
    """
    Best-effort proxy mapping.
    Supports:
      proxySettings: { "proxyUrls": ["http://user:pass@host:port"] }
      proxySettings: { "url": "http://..." }
    If you use Apify Proxy on platform, you usually won't need this locally.
    """
    ps = config.get("proxySettings") or {}
    if not isinstance(ps, dict):
        return None

    url = ps.get("url")
    if isinstance(url, str) and url.strip():
        return {"server": url.strip()}

    urls = ps.get("proxyUrls")
    if isinstance(urls, list) and urls:
        u = str(urls[0]).strip()
        if u:
            return {"server": u}

    # If user passed Apify-style { useApifyProxy: true }, ignore locally
    return None


async def _new_page(config: Dict[str, Any]) -> Tuple[Any, Browser, BrowserContext, Page]:
    """
    Returns: (playwright, browser, context, page)
    Caller must close in finally.
    """
    headless = bool(config.get("headless", True))
    proxy = _proxy_for_playwright(config)

    playwright = await async_playwright().start()

    launch_kwargs: Dict[str, Any] = {"headless": headless}
    if proxy:
        launch_kwargs["proxy"] = proxy

    browser = await playwright.chromium.launch(**launch_kwargs)

    context = await browser.new_context(
        user_agent=(
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36"
        )
    )
    page = await context.new_page()
    page.set_default_timeout(45000)
    return playwright, browser, context, page


# -----------------------------
# REAL APPS FLOW (MINIMAL + XLSX WRAP)
# -----------------------------
# -----------------------------
# REAL APPS FLOW (MINIMAL + XLSX WRAP + FREEZE + FILTER)
# -----------------------------
async def run_apps_flow(config: Dict[str, Any], project_root: str) -> None:
    Actor.log.info("APPS_FLOW_FILE=%s", __file__)
    Actor.log.info("✅ USING apps_flow.py (APPS v2 minimal + xlsx freeze/filter)")

    dataset = await Actor.open_dataset()
    project_root_path = Path(project_root)

    category_group = (config.get("categoryGroup") or "business-needs").strip().lower().replace("_", "-")
    sphere = (config.get("sphere") or category_group).strip().lower().replace("_", "-")

    preset_list = _as_list_str(config.get("categoryPreset"))
    category_preset = (preset_list[0].strip().lower() if preset_list else category_group)

    # ✅ use appGroup first (new input), fallback to categoryPreset (legacy)
    app_groups = _as_list_str(config.get("appGroup"))
    if not app_groups:
        app_groups = _as_list_str(config.get("categoryPreset"))
    app_groups = [s.strip().lower() for s in app_groups if str(s).strip()]

    if not app_groups:
        raise ValueError("APPS flow requires appGroup[] (preferred) or categoryPreset[] (legacy).")

    max_pages = _to_int(config.get("maxPages"), 2)

    # Optional filters (pricingFilter should be a string-select, but tolerate list)
    pf_raw = config.get("pricingFilter")
    if isinstance(pf_raw, list):
        pf = (pf_raw[0] if pf_raw else "")
    else:
        pf = (pf_raw or "")
    pf = str(pf).strip().lower()
    pricing_filter = [pf] if pf else []

    min_rating = _to_float(config.get("minRating"), 0.0)
    min_rating = max(0.0, min(5.0, min_rating))

    Actor.log.info(
        "[APPS] Inputs: categoryGroup=%s sphere=%s categoryPreset=%s appGroup=%s maxPages=%s",
        category_group, sphere, category_preset, app_groups, max_pages
    )

    records: List[Dict[str, Any]] = []

    async with async_playwright() as p:
        browser: Browser = await p.chromium.launch(headless=bool(config.get("headless", True)))
        context: BrowserContext = await browser.new_context()
        page: Page = await context.new_page()

        try:
            for app_group in app_groups:
                explore_url = build_explore_url(category_group, app_group)
                Actor.log.info("[APPS] Discovering: appGroup=%s -> %s", app_group, explore_url)

                app_urls = await discover_app_urls(
                    page,
                    category_group=category_group,
                    sphere_or_category=app_group,
                    max_pages=max_pages,
                )
                Actor.log.info("[APPS] Discovered %s app URLs for appGroup=%s", len(app_urls), app_group)

                for u in app_urls:
                    try:
                        detail = await extract_app_detail(page, u)

                        rec: Dict[str, Any] = {
                            "categoryGroup": category_group,
                            "sphere": sphere,
                            "categoryPreset": category_preset,
                            "appGroup": app_group,

                            "colun_name": detail.get("app_name") or "",
                            "app_url": detail.get("app_url") or u,

                            "short_description": detail.get("short_description") or "",

                            "rating": detail.get("rating"),
                            "reviews": detail.get("reviews_count"),
                            "price": detail.get("price_text") or "",

                            "last_seen": datetime.utcnow().isoformat() + "Z",
                        }

                        records.append(rec)

                    except Exception as e:
                        Actor.log.warning("[APPS] Detail failed url=%s err=%s", u, e)

        finally:
            await context.close()
            await browser.close()

    Actor.log.info("[APPS] Total collected records (pre-filter): %s", len(records))

    # Filters
    filtered = records

    if min_rating > 0:
        before = len(filtered)
        filtered = [r for r in filtered if _to_float(r.get("rating"), 0.0) >= min_rating]
        Actor.log.info("[APPS] minRating=%s filtered %s -> %s", min_rating, before, len(filtered))

    if pricing_filter:
        before = len(filtered)
        pfset = set(pricing_filter)

        def _match_price_bucket(p: str) -> str:
            t = (p or "").lower()
            if "free" in t:
                return "free"
            if "trial" in t or "freemium" in t:
                return "freemium"
            if "$" in t or "per " in t or "subscription" in t:
                return "paid"
            return "unknown"

        filtered = [r for r in filtered if _match_price_bucket(r.get("price", "")) in pfset]
        Actor.log.info("[APPS] pricingFilter=%s filtered %s -> %s", pricing_filter, before, len(filtered))

    if not filtered:
        Actor.log.warning("[APPS] No records after filters. Nothing to export.")
        return

    # Push dataset
    for rec in filtered:
        await dataset.push_data(rec)
    Actor.log.info("[APPS] Pushed %s records to default dataset", len(filtered))

    # Exports (KV)
    df = pd.DataFrame(filtered)
    await Actor.set_value("APPS.csv", df.to_csv(index=False), content_type="text/csv")

    # XLSX wrap + freeze + filter
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    df.to_excel(buf, index=False)
    buf.seek(0)

    wb = load_workbook(buf)
    ws = wb.active

    for cell in ws[1]:
        cell.font = Font(bold=True)

    wrap = Alignment(wrap_text=True, vertical="top")

    headers: Dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        headers[str(cell.value).strip()] = col_idx

    def _apply_wrap(header_name: str, width: int) -> None:
        col_idx = headers.get(header_name)
        if not col_idx:
            return
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col_idx).alignment = wrap

    _apply_wrap("colun_name", 38)
    _apply_wrap("short_description", 70)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    out = io.BytesIO()
    wb.save(out)

    await Actor.set_value(
        "APPS.xlsx",
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # Optional local copies
    try:
        df.to_csv(project_root_path / "APPS.csv", index=False)
        df.to_excel(project_root_path / "APPS.xlsx", index=False)
    except Exception:
        pass

    Actor.log.info("[APPS] Export finished ✅")