from __future__ import annotations

import io
import math
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd
from apify import Actor
from playwright.async_api import async_playwright, Browser, BrowserContext, Page

import salesforce_appexchange_engine.scraper_apps as sa

from salesforce_appexchange_engine.scraper_apps import (
    build_explore_url,
    discover_app_urls,
    extract_app_detail,
)
from src.mi.report import generate_and_save_reports
from apify_client.errors import ApifyApiError


DATASET_FIELDS = [
    "listing_id",
    "categoryGroup",
    "sphere",
    "categoryPreset",
    "appGroup",
    "app_name",
    "name",
    "app_url",
    "short_description",
    "pricing_model",
    "price",
    "rating",
    "rating_bucket",
    "reviews",
    "reviews_bucket",
    "market_segment",
    "last_seen",
    "last_seen_at",
    "category_preset",
]


def _coerce_str(v: Any) -> str:
    return str(v).strip() if v is not None else ""


def _coerce_float_or_none(v: Any):
    if v is None or v == "":
        return None
    try:
        x = float(v)
        if math.isnan(x) or math.isinf(x):
            return None
        return x
    except Exception:
        return None


def _coerce_int_or_none(v: Any):
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except Exception:
        return None


def _shape_for_dataset(rec: Dict[str, Any]) -> Dict[str, Any]:
    r = dict(rec)
    r = _apply_schema_compat(r)

    out: Dict[str, Any] = {k: None for k in DATASET_FIELDS}

    for k in [
        "listing_id", "categoryGroup", "sphere", "categoryPreset", "appGroup",
        "app_name", "name", "app_url", "short_description", "pricing_model",
        "price", "rating_bucket", "reviews_bucket", "market_segment",
        "last_seen", "last_seen_at", "category_preset"
    ]:
        if k in out:
            out[k] = _coerce_str(r.get(k))

    out["rating"] = _coerce_float_or_none(r.get("rating"))
    out["reviews"] = _coerce_int_or_none(r.get("reviews"))

    return out


def _as_list_str(v: Any, *, lower: bool = True) -> List[str]:
    if v is None:
        return []
    if isinstance(v, str):
        v = [v]
    if not isinstance(v, list):
        return []
    out: List[str] = []
    for x in v:
        s = str(x).strip()
        if not s:
            continue
        out.append(s.lower() if lower else s)
    return out


def _to_int(v: Any, default: int) -> int:
    try:
        if v is None or v == "":
            return default
        return int(float(v))
    except Exception:
        return default


def _to_float(v: Any, default: float = 0.0) -> float:
    try:
        if v is None or v == "":
            return default
        return float(v)
    except Exception:
        return default


def _rating_bucket(r: Any) -> str:
    x = _to_float(r, 0.0)
    if x >= 4.5:
        return "4.5+"
    if x >= 4.0:
        return "4.0-4.49"
    if x >= 3.5:
        return "3.5-3.99"
    if x > 0:
        return "<3.5"
    return "unrated"


def _reviews_bucket(n: Any) -> str:
    x = _to_int(n, 0)
    if x >= 1000:
        return "1000+"
    if x >= 250:
        return "250-999"
    if x >= 50:
        return "50-249"
    if x >= 10:
        return "10-49"
    if x >= 1:
        return "1-9"
    return "0"


def _pricing_model_from_text(price_text: str) -> str:
    t = (price_text or "").lower()

    if "nonprofit" in t or "non-profit" in t or "discount" in t:
        return "nonprofit-discount"

    if "trial" in t or "try it free" in t or "free trial" in t or "freemium" in t:
        return "freemium"

    if "$" in t or "/month" in t or "per user" in t or "subscription" in t:
        return "paid"

    if "free" in t:
        return "free"

    return "unknown"


def _market_segment(category_group: str, category_preset: str, app_group: str) -> str:
    cg = (category_group or "").strip().lower()
    cp = (category_preset or "").strip().lower()
    ag = (app_group or "").strip().lower()

    if cg == "business-needs":
        return f"business_needs / {ag}"
    if cg == "industries":
        return f"industries / {cp or ag}"
    if cg == "products":
        return f"products / {cp or ag}"
    return f"{cg or 'unknown'} / {ag or cp or 'unknown'}"


def _pick_app_groups(config: Dict[str, Any], *, category_group: str) -> List[str]:
    explicit = _as_list_str(config.get("appGroup"), lower=False)
    explicit = [s.strip() for s in explicit if s.strip()]
    if explicit:
        return explicit

    resolved = config.get("resolvedAppTypes") or []

    if category_group == "products":
        if isinstance(resolved, list) and resolved:
            titles: List[str] = []
            for t in resolved:
                if isinstance(t, dict):
                    title = str(t.get("title") or "").strip()
                    if title:
                        titles.append(title)
            if titles:
                return titles

        app_types = _as_list_str(config.get("appTypes"), lower=False)
        app_types = [a.strip() for a in app_types if a.strip()]
        if app_types:
            return [a.replace("-", " ") for a in app_types]

    preset = _as_list_str(config.get("categoryPreset"), lower=False)
    preset = [s.strip() for s in preset if s.strip()]
    if preset:
        return preset

    return []


def _json_safe(v: Any):
    if v is None:
        return None

    if isinstance(v, (str, int, bool)):
        return v

    if isinstance(v, float):
        if math.isnan(v) or math.isinf(v):
            return None
        return v

    if isinstance(v, Decimal):
        return float(v)

    if isinstance(v, datetime):
        return v.isoformat()

    if isinstance(v, dict):
        return {str(k): _json_safe(vv) for k, vv in v.items()}

    if isinstance(v, (list, tuple)):
        return [_json_safe(x) for x in v]

    if isinstance(v, set):
        return [_json_safe(x) for x in sorted(list(v), key=lambda x: str(x))]

    return str(v)


def _apply_schema_compat(rec: Dict[str, Any]) -> Dict[str, Any]:
    u = rec.get("app_url") or ""

    if not rec.get("listing_id") and "listingId=" in u:
        rec["listing_id"] = u.split("listingId=", 1)[1].split("&", 1)[0].strip() or None

    if rec.get("last_seen") and not rec.get("last_seen_at"):
        rec["last_seen_at"] = rec.get("last_seen")

    if rec.get("categoryPreset") and not rec.get("category_preset"):
        rec["category_preset"] = rec.get("categoryPreset")

    if rec.get("app_name") and not rec.get("name"):
        rec["name"] = rec.get("app_name")

    return rec


async def run_apps_flow(config: Dict[str, Any], project_root: str) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    Actor.log.info("✅ USING apps_flow.py (clean, Store-safe)")
    Actor.log.info("[debug] scraper_apps path = %s", sa.__file__)

    dataset = await Actor.open_dataset()
    _ = Path(project_root)

    category_group = (config.get("categoryGroup") or "business-needs").strip().lower().replace("_", "-")
    sphere = (config.get("sphere") or category_group).strip().lower().replace("_", "-")

    preset_list = _as_list_str(config.get("categoryPreset"), lower=False)
    category_preset = (preset_list[0].strip() if preset_list else category_group)

    app_groups = _pick_app_groups(config, category_group=category_group)
    if not app_groups:
        raise ValueError("APPS flow requires appGroup[] or resolvedAppTypes/categoryPreset.")

    max_pages = _to_int(config.get("maxPages"), 2)

    pf_raw = config.get("pricingFilter")
    if isinstance(pf_raw, list):
        pf = (pf_raw[0] if pf_raw else "")
    else:
        pf = (pf_raw or "")
    pf = str(pf).strip().lower()
    allowed_pf = {"free", "freemium", "paid", "nonprofit-discount", "unknown"}
    pricing_filter = [pf] if pf in allowed_pf else []
    if pf and not pricing_filter:
        Actor.log.warning("[APPS] pricingFilter='%s' not recognized; ignoring. allowed=%s", pf, sorted(allowed_pf))

    min_rating = max(0.0, min(5.0, _to_float(config.get("minRating"), 0.0)))

    Actor.log.info(
        "[APPS] Inputs: categoryGroup=%s sphere=%s categoryPreset=%s appGroups=%s maxPages=%s minRating=%s pricingFilter=%s",
        category_group,
        sphere,
        category_preset,
        app_groups,
        max_pages,
        min_rating,
        pricing_filter,
    )

    records: List[Dict[str, Any]] = []

    async with async_playwright() as p:
        browser: Browser = await p.chromium.launch(headless=bool(config.get("headless", True)))
        context: BrowserContext = await browser.new_context()
        page: Page = await context.new_page()

        try:
            for app_group in app_groups:
                explore_url = build_explore_url(category_group, app_group)

                app_urls = await discover_app_urls(
                    page,
                    category_group=category_group,
                    sphere_or_category=app_group,
                    max_pages=max_pages,
                )

                Actor.log.info(
                    "[APPS] Discovered %s app URLs for appGroup=%s (explore=%s)",
                    len(app_urls),
                    app_group,
                    explore_url,
                )

                for u in app_urls:
                    try:
                        detail = await extract_app_detail(page, u)

                        app_name = (detail.get("app_name") or "").strip()
                        app_url = (detail.get("app_url") or u).strip()
                        short_desc = (detail.get("short_description") or "").strip()

                        rating = detail.get("rating")
                        reviews = detail.get("reviews_count")
                        price_text = (detail.get("price_text") or "").strip()

                        pricing_model = _pricing_model_from_text(price_text)

                        rec: Dict[str, Any] = {
                            "categoryGroup": category_group,
                            "sphere": sphere,
                            "categoryPreset": category_preset,
                            "appGroup": app_group,
                            "app_name": app_name,
                            "app_url": app_url,
                            "short_description": short_desc,
                            "pricing_model": pricing_model,
                            "rating_bucket": _rating_bucket(rating),
                            "reviews_bucket": _reviews_bucket(reviews),
                            "market_segment": _market_segment(category_group, category_preset, app_group),
                            "rating": rating,
                            "reviews": reviews,
                            "price": price_text,
                            "last_seen": datetime.utcnow().isoformat() + "Z",
                        }

                        records.append(rec)

                    except Exception as e:
                        Actor.log.warning("[APPS] Detail failed url=%s err=%s", u, e)

        finally:
            await context.close()
            await browser.close()

    Actor.log.info("[APPS] Total collected records (pre-filter): %s", len(records))

    filtered = records

    if min_rating > 0:
        before = len(filtered)

        def _keep(r: Dict[str, Any]) -> bool:
            rv = r.get("rating")
            if rv is None or rv == "":
                return True
            return _to_float(rv, 0.0) >= min_rating

        filtered = [r for r in filtered if _keep(r)]
        Actor.log.info("[APPS] minRating=%s filtered %s -> %s (kept unrated)", min_rating, before, len(filtered))

    if pricing_filter:
        before = len(filtered)
        pfset = {str(x).strip().lower() for x in pricing_filter if str(x).strip()}
        filtered2 = [r for r in filtered if str(r.get("pricing_model", "unknown")).lower() in pfset]
        if not filtered2:
            Actor.log.warning("[APPS] pricingFilter would drop everything; ignoring for this run. pf=%s", list(pfset))
        else:
            filtered = filtered2
            Actor.log.info("[APPS] pricingFilter=%s filtered %s -> %s", list(pfset), before, len(filtered))

    export_cols = [
        "categoryGroup",
        "sphere",
        "categoryPreset",
        "appGroup",
        "app_name",
        "app_url",
        "short_description",
        "pricing_model",
        "price",
        "rating",
        "rating_bucket",
        "reviews",
        "reviews_bucket",
        "market_segment",
        "last_seen",
    ]

    if not filtered:
        Actor.log.warning("[APPS] No records after filters. Exporting EMPTY APPS.csv/APPS.xlsx to KV.")
        df_empty = pd.DataFrame(columns=export_cols)
        await Actor.set_value("APPS.csv", df_empty.to_csv(index=False), content_type="text/csv")

        buf_empty = io.BytesIO()
        df_empty.to_excel(buf_empty, index=False)
        await Actor.set_value(
            "APPS.xlsx",
            buf_empty.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        return

    pushed = 0
    for rec in filtered:
        try:
            shaped = _shape_for_dataset(rec)
            await dataset.push_data(shaped)
            pushed += 1
        except ApifyApiError as e:
            Actor.log.error("[DATASET SCHEMA FAIL] url=%s listing_id=%s", rec.get("app_url"), rec.get("listing_id"))
            try:
                Actor.log.error("[DATASET SCHEMA FAIL] message=%s", getattr(e, "message", str(e)))
                Actor.log.error("[DATASET SCHEMA FAIL] details=%s", getattr(e, "data", None))
            except Exception:
                pass
            raise
        except Exception as e:
            Actor.log.error("[DATASET PUSH FAIL] url=%s err=%s", rec.get("app_url"), e)
            raise

    Actor.log.info("[APPS] Pushed %s records to default dataset", pushed)

    df = pd.DataFrame(filtered).reindex(columns=export_cols)
    await Actor.set_value("APPS.csv", df.to_csv(index=False), content_type="text/csv")

    buf = io.BytesIO()
    df.to_excel(buf, index=False)
    buf.seek(0)

    wb = load_workbook(buf)
    ws = wb.active
    ws.freeze_panes = "A2"
    last_row = ws.max_row
    last_col = ws.max_column
    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"

    wrap = Alignment(wrap_text=True, vertical="top")

    header_to_col: Dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        header_to_col[(str(cell.value or "")).strip()] = col_idx

    def _wrap_col(header_name: str, width: int) -> None:
        col_idx = header_to_col.get(header_name)
        if not col_idx:
            return
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col_idx).alignment = wrap

    def _set_width(header_name: str, width: int) -> None:
        col_idx = header_to_col.get(header_name)
        if not col_idx:
            return
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    _wrap_col("app_name", 38)
    _wrap_col("short_description", 70)
    _wrap_col("price", 45)

    _set_width("app_url", 55)
    _set_width("pricing_model", 16)
    _set_width("market_segment", 24)

    out = io.BytesIO()
    wb.save(out)

    await Actor.set_value(
        "APPS.xlsx",
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    Actor.log.info("[APPS] Export finished ✅")
    # ---------------------------------
    # Market Intelligence / report files
    # ---------------------------------
    analysis_options = config.get("analysisOptions") or {}

    try:
        mi_result = await generate_and_save_reports(
            df=df,
            config=config,
            analysis_options=analysis_options,
            kv_prefix=str(analysis_options.get("reportKvKeyPrefix") or ""),
        )
        Actor.log.info("[APPS] MI/report generation finished: %s", mi_result)
    except Exception as e:
        Actor.log.exception("[APPS] MI/report generation failed: %s", e)